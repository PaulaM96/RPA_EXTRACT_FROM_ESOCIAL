import os
import logging
import datetime
import pyodbc
from openpyxl import Workbook, load_workbook

# Configuração do log
logging.basicConfig(level=logging.INFO, filename="programa.log", format="%(asctime)s - %(levelname)s - %(message)s")

def inicializar_planilha(caminho_planilha):
    diretorio = os.path.dirname(caminho_planilha)
    if not os.path.exists(diretorio):
        os.makedirs(diretorio)
    if not os.path.exists(caminho_planilha):
        wb = Workbook()
        ws = wb.active
        ws.append(['Centro de Custo', 'CPF', 'Nome'])
        wb.save(caminho_planilha)
    else:
        wb = load_workbook(caminho_planilha)
    return wb

def salvar_dados_excel(caminho_planilha, dados):
    wb = load_workbook(caminho_planilha)
    ws = wb.active
    for dado in dados:
        ws.append(dado)
    wb.save(caminho_planilha)

def conectar_banco_dados():
    try:
        conexao = pyodbc.connect(
            'DRIVER={driver};'
            'SERVER=server;'
            'DATABASE=db;'
            'UID=user;'
            'PWD=password;'
        )
        return conexao
    except pyodbc.Error as erro:
        logging.error(f"Erro ao conectar ao banco de dados: {erro}")
        return None

def buscar_funcionarios_ra_cic(conexao, cc):
    funcionarios_ra_cic = []
    if conexao:
        try:
            cursor = conexao.cursor()
            data_atual = datetime.datetime.now()
            primeiro_dia_mes_anterior = (data_atual.replace(day=1) - datetime.timedelta(days=1)).replace(day=1)
            ultimo_dia_mes_anterior = data_atual.replace(day=1) - datetime.timedelta(days=1)
            data_inicio = primeiro_dia_mes_anterior.strftime('%Y%m%d')
            data_fim = ultimo_dia_mes_anterior.strftime('%Y%m%d')
            consulta = f"""
                SELECT RA_CIC 
                FROM SRA010 
                WHERE RA_CC ='{cc}' 
                AND (RA_DEMISSA = '' OR RA_DEMISSA >'{data_fim}')
                AND (RA_ADMISSA <= '{data_fim}')
                AND ((RA_MAT NOT IN (select RE_MATD FROM SRE010 WHERE (RE_DATA BETWEEN '{data_inicio}' AND '{data_fim}') AND (RE_CCD='{cc}') AND D_E_L_E_T_=''))
                OR (RA_MAT IN (select RE_MATD FROM SRE010 WHERE (RE_DATA BETWEEN '{data_inicio}' AND '{data_fim}') AND (RE_CCP='{cc}') AND D_E_L_E_T_='')))
            """
            cursor.execute(consulta)
            for row in cursor.fetchall():
                ra_cic = row[0]
                funcionarios_ra_cic.append(ra_cic)
            cursor.close()
        except pyodbc.Error as erro:
            logging.error(f"Erro ao buscar funcionários no banco de dados: {erro}")
    else:
        logging.error("Falha na conexão ao banco de dados.")
    return funcionarios_ra_cic

def buscar_nome_completo(conexao, ra_cic):
    try:
        cursor = conexao.cursor()
        consulta = f"SELECT RA_NOMECMP FROM SRA010 WHERE RA_CIC = '{ra_cic}'"
        cursor.execute(consulta)
        resultado = cursor.fetchone()
        cursor.close()
        if resultado:
            return resultado[0]
        else:
            logging.error(f"Nome completo não encontrado para RA_CIC: {ra_cic}")
            return None
    except pyodbc.Error as erro:
        logging.error(f"Erro ao buscar o nome completo do funcionário: {ra_cic}, {erro}")
        return None

def processar_centro_custo(conexao, cc, caminho_planilha):
    funcionarios = buscar_funcionarios_ra_cic(conexao, cc)
    dados = []
    for ra_cic in funcionarios:
        nome_completo = buscar_nome_completo(conexao, ra_cic)
        if nome_completo:
            dados.append([cc, ra_cic, nome_completo])
    salvar_dados_excel(caminho_planilha, dados)

def main():
    #
    ccs = ['cc1','cc2']
    caminho_planilha = r"\\servidor\caminho\entradaCP.xlsx"

    # Inicializar a planilha
    inicializar_planilha(caminho_planilha)

    conexao = conectar_banco_dados()
    if conexao:
        for cc in ccs:
            processar_centro_custo(conexao, cc, caminho_planilha)
        conexao.close()
    else:
        logging.error("Não foi possível conectar ao banco de dados.")

if __name__ == "__main__":
    main()
