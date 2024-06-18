import os
import logging
from seleniumbase import SB
from openpyxl import load_workbook, Workbook

# Configuração do log
logging.basicConfig(level=logging.INFO, filename="programa.log", format="%(asctime)s - %(levelname)s - %(message)s")

def inicializar_planilha(caminho_planilha):
    diretorio = os.path.dirname(caminho_planilha)
    if not os.path.exists(diretorio):
        os.makedirs(diretorio)
    if not os.path.exists(caminho_planilha):
        wb = Workbook()
        ws = wb.active
        ws.append(['Centro de Custo', 'CPF', 'Nome', 'URL de Download', 'Tipo'])
        wb.save(caminho_planilha)
    else:
        wb = load_workbook(caminho_planilha)
    return wb

def obter_dados_planilha(caminho_planilha, lote):
    wb = load_workbook(caminho_planilha)
    ws = wb.active
    cpfs = []
    for row in ws.iter_rows(min_row=2, max_row=lote+1, values_only=True):
        if row and len(row) >= 4:
            cc, cpf, nome, url = row[:4]
            if None in (cc, cpf, nome, url):
                logging.error("Registro inválido encontrado: Centro de Custo, CPF ou Nome está vazio.")
                continue
            cpfs.append((cc, cpf, nome.strip(), url.strip()))  # Removendo espaços extras
    return cpfs

def salvar_arquivo(sb, url, nome_funcionario, caminho_arquivo):
    # Sanitizar o nome do arquivo para evitar caracteres inválidos
    nome_funcionario = ''.join(e for e in nome_funcionario if e.isalnum() or e.isspace()).strip()
    nome_arquivo = f"{nome_funcionario}"
    caminho_arquivo = os.path.join(caminho_arquivo, nome_arquivo)
    sb.execute_script(f"window.open('{url}', 'newtab');")
    sb.execute_cdp_cmd("Page.setDownloadBehavior", {"behavior": "allow", "downloadPath": caminho_arquivo})
    sb.sleep(5)  # Ajuste conforme necessário para esperar o download concluir

    # Aguarde até que o arquivo seja baixado
    while not os.path.exists(caminho_arquivo):
        sb.sleep(1)

    # Renomear o arquivo para o nome do funcionário
    arquivo_baixado = os.path.join(caminho_arquivo, os.path.basename(url))
    if os.path.exists(arquivo_baixado):
        os.rename(arquivo_baixado, caminho_arquivo)
    return caminho_arquivo

def remover_cpfs_da_planilha(caminho_planilha, cpfs):
    wb = load_workbook(caminho_planilha)
    ws = wb.active
    for cpf in cpfs:
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[1].value == cpf:
                ws.delete_rows(row[0].row, 1)
                break
    wb.save(caminho_planilha)

def processar_downloads_em_lotes(caminho_planilha, lote):
    cpfs = obter_dados_planilha(caminho_planilha, lote)
    
    if not cpfs:
        logging.info("Nenhum CPF encontrado na planilha de entrada.")
        return False
    
    resultados = []
    cpfs_processados = []
    
    with SB(uc=True) as sb:
        realizar_login(sb)
        for cc, cpf, nome, url in cpfs:
            if None in (cc, cpf, nome, url):
                continue
            try:
                logging.info(f"Processando download para CPF: {cpf}, Nome: {nome}, Centro de Custo: {cc}")
                caminho_pasta = os.path.join(r"\\servidor\caminho\CP", cc)
                if not os.path.exists(caminho_pasta):
                    os.makedirs(caminho_pasta)
                
                caminho_arquivo = salvar_arquivo(sb, url, nome, caminho_pasta)
                if os.path.exists(caminho_arquivo):
                    logging.info(f"Download concluído e salvo em: {caminho_arquivo}")
                    resultados.append((cc, cpf, nome, url))
                    cpfs_processados.append(cpf)
                else:
                    logging.error(f"Erro no download para CPF: {cpf}, Nome: {nome}, Centro de Custo: {cc}")
            except Exception as e:
                logging.error(f"Erro ao processar download para CPF: {cpf}, Nome: {nome}, Centro de Custo: {cc}, Erro: {e}")
                continue
        
        if cpfs_processados:
            remover_cpfs_da_planilha(caminho_planilha, cpfs_processados)
    
    return resultados

def realizar_login(sb):
    sb.get("chrome://settings/security")
    sb.sleep(10)
    url_esocial = "https://login.esocial.gov.br/login.aspx"
    sb.get(url_esocial)
    sb.wait_for_element_present(".br-button.sign-in.large")
    sb.click(".br-button.sign-in.large")
    sb.wait_for_element_present("#login-certificate")
    sb.click("#login-certificate")
    sb.sleep(30)

def main():
    caminho_planilha = r"\\servidor\caminho\saidaCP.xlsx"
    inicializar_planilha(caminho_planilha)
    lote = 30  # Número de registros por lote
    
    while True:
        resultados = processar_downloads_em_lotes(caminho_planilha, lote)
        if not resultados:
            break

if __name__ == "__main__":
    main()
