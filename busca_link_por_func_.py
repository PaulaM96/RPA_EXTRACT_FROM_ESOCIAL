import os
import logging
import datetime
import time
from openpyxl import Workbook, load_workbook
from seleniumbase import SB
from selenium.webdriver.common.keys import Keys

# Configuração do log
logging.basicConfig(level=logging.INFO, filename="programa.log", format="%(asctime)s - %(levelname)s - %(message)s")

def inicializar_planilha(caminho_planilha, colunas):
    diretorio = os.path.dirname(caminho_planilha)
    if not os.path.exists(diretorio):
        os.makedirs(diretorio)
    if not os.path.exists(caminho_planilha):
        wb = Workbook()
        ws = wb.active
        ws.append(colunas)
        wb.save(caminho_planilha)
    else:
        wb = load_workbook(caminho_planilha)
    return wb

def salvar_dados_excel(caminho_planilha, dados):
    wb = load_workbook(caminho_planilha)
    ws = wb.active
    for dado in dados:
        ws.append(dado)
    wb.save(caminho_planilha)

def ler_cpfs_da_planilha(caminho_planilha, lote=10):
    try:
        wb = load_workbook(caminho_planilha)
        ws = wb.active
        cpfs = []
        for row in ws.iter_rows(min_row=2, max_row=lote+1, values_only=True):
            if row and len(row) >= 3:
                cc, cpf, nome = row[:3]
                if None in (cc, cpf, nome):
                    logging.error("Registro inválido encontrado: Centro de Custo, CPF ou Nome está vazio.")
                    cpfs.append((cpf,nome,cc))
                    return cpfs  # Retornar None se qualquer campo estiver vazio
                cpfs.append((cpf, nome, cc))
            else:
                logging.warning("Erro na leitura da linha da planilha: Linha vazia ou incompleta.")
        if not cpfs:
            logging.warning("Nenhum CPF encontrado na planilha de entrada.")
        else:
            logging.info(f"{len(cpfs)} CPFs lidos com sucesso da planilha.")
        return cpfs
    except Exception as e:
        logging.error(f"Erro ao ler CPFs da planilha: {e}")
        return None

def remover_cpfs_da_planilha(caminho_planilha, cpfs):
    wb = load_workbook(caminho_planilha)
    ws = wb.active
    for cpf in cpfs:
        for row in ws.iter_rows(min_row=2, values_only=False):
            if row[1].value == cpf:  # Verifica se a célula contém o CPF
                ws.delete_rows(row[0].row, 1)
                break
    wb.save(caminho_planilha)

def buscar_informacoes_esocial(sb, cpf):
    try:
        sb.wait_for_element_present("#PeriodoApuracaoPesquisa")
        input_field = sb.find_element("#PeriodoApuracaoPesquisa")
        input_field.clear()
        sb.sleep(1)
        input_field.send_keys(mes_ano_pesquisa)
        sb.wait_for_element_present("#CpfPesquisa")
        input_field = sb.find_element("#CpfPesquisa")
        input_field.clear()
        sb.sleep(1)
        input_field.send_keys(cpf)
        input_field.send_keys(Keys.ENTER)
        sb.wait_for_ready_state_complete()
        if sb.is_element_visible("p:contains('Nenhum registro encontrado')"):
            logging.warning(f"Nenhum registro encontrado para o CPF: {cpf}")
            return None
        download_link = sb.find_element("a.btn.btn-primary.pull-right.margin-right-10px.download").get_attribute("href")
        return download_link
    except Exception as e:
        logging.error(f"Erro ao buscar informações no eSocial para CPF {cpf}: {e}")
        return None

def processar_lote(caminho_entrada, caminho_saida, mes_ano_pesquisa):
    cpfs = ler_cpfs_da_planilha(caminho_entrada)
    
    if not cpfs:
        logging.info("Nenhum CPF encontrado na planilha de entrada.")
        return False
    
    resultados = []
    cpfs_processados = []
    
    with SB(uc=True) as sb:
        url_esocial = "https://login.esocial.gov.br/login.aspx"
        sb.get(url_esocial)
        sb.wait_for_element_present(".br-button.sign-in.large")
        sb.click(".br-button.sign-in.large")
        sb.wait_for_element_present("#login-certificate")
        sb.click("#login-certificate")
        sb.wait_for_ready_state_complete(timeout=40)
        sb.wait_for_element_present("#menuFolhaPagamento")
        sb.click('#menuFolhaPagamento')
        sb.wait_for_element_present("#menuTotalizadores")
        sb.click('#menuTotalizadores')
        sb.wait_for_element_present("#menuTotalizadoresTrabalhador")
        sb.execute_script("document.querySelector('#menuTotalizadoresTrabalhador').click()")
        sb.wait_for_element_present("#menuContribuicaoSocialTrabalhador")
        sb.execute_script("document.querySelector('#menuContribuicaoSocialTrabalhador').click()")
            
        for cpf, nome, cc in cpfs:  # Corrigir a iteração sobre a lista de cpfs
            if None in (cc, cpf, nome):
                sb.sleep(10)
                salvar_dados_excel(caminho_saida, resultados)
                remover_cpfs_da_planilha(caminho_entrada, cpfs_processados)
                return False
            else:
                download_link = buscar_informacoes_esocial(sb, cpf)
                if download_link:
                    resultados.append([cc, cpf, nome, download_link])
                    cpfs_processados.append(cpf)
    
    if resultados:
        salvar_dados_excel(caminho_saida, resultados)
        remover_cpfs_da_planilha(caminho_entrada, cpfs_processados)
    
    return True

def main():
    
    caminho_entrada = r"\\servidor\caminho\entradaCP.xlsx"
    caminho_saida = r"\\servidor\caminho\saidaCP.xlsx"
    inicializar_planilha(caminho_entrada, ['Centro de Custo', 'CPF', 'Nome'])
    inicializar_planilha(caminho_saida, ['Centro de Custo', 'CPF', 'Nome', 'Link de Download'])
    while True:
        if not processar_lote(caminho_entrada, caminho_saida, mes_ano_pesquisa):
            break
        time.sleep(1)  # Pequeno intervalo entre os lotes para evitar sobrecarga

if __name__ == "__main__":
    data_atual = datetime.datetime.now()
    mes_ano_pesquisa = (data_atual - datetime.timedelta(days=35)).strftime("%m%Y")
    main()
